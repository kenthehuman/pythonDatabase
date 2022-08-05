"""
To extract parts and stuff from excel sheet and add it into a database
This script extracts from the DC202 BOM made 7/14/2022, to add some purchased parts to the database.
By:Ken
Date:8/1/2022
"""
from openpyxl import load_workbook
from classes import DEpart
from mapping import PARTNAME, PARTDESCRIPTION, PARTNUMBER, PARTSUPPLIER

import sqlite3

workbook = load_workbook("BOM_Sample.xlsx")
sheet = workbook.active

hardware = []

# for value in sheet.iter_rows(min_row=10,
#                             max_row=10,
#                             min_col=2,
#                             max_col=6,
#                             values_only=True):
#     print(value)
# 
# gets the description of the first thing
# hardware[0][0].description
# hardware[0][0].name
# hardware[0][0].part_number
# hardware[0][0].supplier


# convert list to a list of tuple to insert with .execute method
hardware = [tuple(hardware)]



con = sqlite3.connect("bom.db")
cur = con.cursor()

cur.execute()

cur.execute("""insert into part 
            (part, description, part_number, supplier) values (?, ?, ?), 
            ((hardware[0].description,), (hardware[0].name,), 
            (hardware[0].part_number,), (hardware[0].supplier,))""")

# doesnt work
# # gets from rows noted and loads into hardware
# for row in sheet.iter_rows(min_row=13,
#                             max_row=31,
#                             values_only=True):
#     product = DEpart(name=row[NAME],
#                     description=row[DESCRIPTION],
#                     part_number=row[PART_NUMBER],
#                     supplier=row[SUPPLIER])
#     hardware.append(product)
# # puts multiple variables into a query for inserting

query = """insert into parts 
            (partName, partDescription, partNumber, partSupplier) values (?, ?, ?, ?)"""




# Extracted all items from turntable from the excel sheet ('Turntable') DC202BOM
hardware=[]
for row in sheet.iter_rows(min_row=3,
                            max_row=16,
                            values_only=True):
    product = DEpart(partName=row[PARTNAME],
                    partDescription=row[PARTDESCRIPTION],
                    partNumber=row[PARTNUMBER],
                    partSupplier=row[PARTSUPPLIER])
    hardware.append(product)

for row in sheet.iter_rows(min_row=19,
                            max_row=33,
                            values_only=True):
    product = DEpart(partName=row[PARTNAME],
                    partDescription=row[PARTDESCRIPTION],
                    partNumber=row[PARTNUMBER],
                    partSupplier=row[PARTSUPPLIER])
    hardware.append(product)

for i in range(len(hardware)):        
    partName = hardware[i].partName
    partDescription = hardware[i].partDescription
    partNumber = hardware[i].partNumber
    partSupplier = hardware[i].partSupplier
    query_variable = [(partName), (partDescription), (partNumber), (partSupplier)]
    print(query_variable)
    cur.execute(query, query_variable)



